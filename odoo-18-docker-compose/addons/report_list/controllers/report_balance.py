import json
import csv
import io
from datetime import datetime
from odoo import http
from odoo.http import request

class ReportBalanceController(http.Controller):
    @http.route('/report-balance', type='http', auth='user', website=True)
    def report_balance_page(self, **kw):
        """
        Route to render the Report Balance page.
        """
        return request.render('report_list.report_balance_page_template', {})

    @http.route('/report-balance/data', type='json', auth='user', methods=['POST'])
    def get_report_data(self, domain=None, search_values=None, limit=10, offset=0, **kw):
        """
        JSON RPC endpoint to fetch report balance data with advanced search.
        """
        try:
            # Sử dụng portfolio.investment thay vì report.balance
            report_model = request.env['portfolio.investment'].sudo()
            
            # Start with the base domain from filters
            if not domain:
                domain = []

            # Add search values from inline search fields
            if search_values:
                for field, value in search_values.items():
                    if value: # Only add to domain if there is a value
                        # Map frontend field names to actual model fields
                        field_mapping = {
                            'so_tai_khoan': 'user_id.partner_id.name',
                            'nha_dau_tu': 'user_id.name',
                            'quy': 'fund_id.name',
                        }
                        actual_field = field_mapping.get(field, field)
                        domain.append((actual_field, 'ilike', value))

            # Resolve pagination params (page -> offset)
            page = int(kw.get('page', 1) or 1)
            limit = int(limit or 10)
            offset = int(kw.get('offset', (page - 1) * limit))

            # Debug: Log domain and parameters
            print(f"Report Balance Domain: {domain}")
            print(f"Limit: {limit}, Offset: {offset}, Page: {page}")
            
            # Build domain from filters (fund, dateFrom, dateTo)
            filters = kw.get('filters') or {}
            fund_filter = filters.get('fund')
            date_from = filters.get('dateFrom') or filters.get('from_date')
            date_to = filters.get('dateTo') or filters.get('to_date')

            if fund_filter:
                try:
                    fund_id_int = int(fund_filter)
                except Exception:
                    fund_id_int = fund_filter
                domain.append(('fund_id', '=', fund_id_int))

            if date_from and date_to and date_from == date_to:
                domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
                domain.append(('create_date', '<=', f"{date_to} 23:59:59"))
            else:
                if date_from:
                    domain.append(('create_date', '>=', f"{date_from} 00:00:00"))
                if date_to:
                    domain.append(('create_date', '<=', f"{date_to} 23:59:59"))

            # Get total count
            total = report_model.search_count(domain)
            
            # Get records with pagination
            records = report_model.search(domain, limit=limit, offset=offset, order='create_date desc, id desc')
            
            # Format data for frontend
            data = []
            for record in records:
                # Get user/partner info
                user_name = record.user_id.name if record.user_id else ''
                partner = record.user_id.partner_id if record.user_id else False
                partner_name = partner.name if partner else ''

                # Fetch số tài khoản ưu tiên từ status.info (investor_profile_management)
                account_number = partner_name
                try:
                    if partner:
                        status_info = request.env['status.info'].sudo().search([('partner_id', '=', partner.id)], limit=1)
                        if status_info and getattr(status_info, 'so_tk', False):
                            account_number = status_info.so_tk
                except Exception:
                    # Fallback giữ nguyên partner_name nếu không có module/record
                    account_number = partner_name

                # Get fund info
                fund_name = record.fund_id.name if record.fund_id else ''
                fund_ticker = record.fund_id.ticker if record.fund_id else ''

                # Get nationality from investor.profile if available
                nationality = ''
                try:
                    if partner:
                        investor_profile = request.env['investor.profile'].sudo().search([('partner_id', '=', partner.id)], limit=1)
                        if investor_profile and investor_profile.nationality:
                            nationality = investor_profile.nationality.name
                except Exception:
                    nationality = ''

                data.append({
                    'id': record.id,
                    'so_tai_khoan': account_number,
                    'so_tk_gdck': '',  # Not available in portfolio.investment
                    'nha_dau_tu': user_name,
                    'so_dien_thoai': record.user_id.partner_id.phone if record.user_id and record.user_id.partner_id else '',
                    'dksh': str(record.amount) if record.amount else '',
                    'email': record.user_id.partner_id.email if record.user_id and record.user_id.partner_id else '',
                    'loai_ndt': 'Trực tiếp',
                    'quoc_tich': nationality,
                    'nvcs': '',  # Not available in portfolio.investment
                    'quy': fund_name,
                    'chuong_trinh': fund_name,
                    'chuong_trinh_ticker': fund_ticker,
                    'ngay_in': record.create_date.strftime('%d/%m/%Y') if getattr(record, 'create_date', None) else '',
                    'so_ccq': record.units or 0,
                    'don_vi': 'VND'
                })
            
            print(f"Returning data: {len(data)} records, total: {total}")
            return {
                'data': data,
                'total': total,
                'page': page,
                'limit': limit
            }
            
        except Exception as e:
            print(f"Error in get_report_data: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'error': str(e),
                'records': [],
                'total': 0,
            }

    @http.route('/report-balance/products', type='json', auth='user', methods=['POST'])
    def get_products(self, **kw):
        """
        JSON RPC endpoint to fetch products for the filter dropdown.
        """
        try:
            # Sử dụng portfolio.investment thay vì report.balance
            products = request.env['portfolio.investment'].sudo().get_products_for_filter()
            print(f"Products found: {products}")
            return products
            
        except Exception as e:
            print(f"Error in get_products: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @http.route('/report-balance/export-pdf', type='http', auth='user', website=True)
    def export_pdf(self, **kw):
        """
        Route to export the report balance data as a PDF.
        """
        try:
            # Sử dụng portfolio.investment thay vì report.balance
            report_model = request.env['portfolio.investment'].sudo()
            
            # Get filters from URL parameters
            selected_product = kw.get('product', '')
            to_date = kw.get('to_date', '')
            loai_ndt = kw.get('loai_ndt', '')
            
            # Build domain based on filters
            domain = []
            if selected_product:
                domain.append(['fund_id.name', '=', selected_product])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            if loai_ndt:
                domain.append(['loai_ndt', '=', loai_ndt])
            
            # Get all records for PDF export (no pagination)
            records = report_model.search(domain, order='id desc')
            
            # Prepare data for template
            data = {
                'records': records,
                'total_records': len(records),
                'selected_product': selected_product,
                'to_date': to_date,
                'loai_ndt': loai_ndt,
                'export_time': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                'company_name': request.env.company.name,
                'user_name': request.env.user.name,
            }
            
            # Generate PDF report
            pdf_content, content_type = request.env['ir.actions.report']._render_qweb_pdf(
                'report_list.report_balance_pdf_template',
                res_ids=[1],  # Pass a dummy ID
                data=data
            )
            
            # Set response headers
            pdf_http_headers = [
                ('Content-Type', 'application/pdf'),
                ('Content-Length', len(pdf_content)),
                ('Content-Disposition', f'attachment; filename="Bao_cao_so_du_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'),
            ]
            
            return request.make_response(pdf_content, headers=pdf_http_headers)
            
        except Exception as e:
            print(f"Error in export_pdf: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất PDF</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-balance/export-xlsx', type='http', auth='user', website=True)
    def export_xlsx(self, **kw):
        """
        Route to export the report balance data as an XLSX.
        """
        try:
            # Sử dụng portfolio.investment thay vì report.balance
            report_model = request.env['portfolio.investment'].sudo()
            
            # Get filters from URL parameters
            selected_product = kw.get('product', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if selected_product:
                domain.append(['fund_id.name', '=', selected_product])
            if from_date:
                domain.append(['create_date', '>=', from_date])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            
            # Get all records for XLSX export (no pagination)
            records = report_model.search(domain, order='id desc')
            
            # Create XLSX content
            import io
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Report Balance"
            
            # Headers
            headers = ['STT', 'Số TK', 'Tên khách hàng', 'Loại NĐT', 'Quốc tịch', 'Số CCQ']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for i, record in enumerate(records, 2):
                # Get user/partner info
                user_name = record.user_id.name if record.user_id else ''
                partner = record.user_id.partner_id if record.user_id else False
                partner_name = partner.name if partner else ''

                # Fetch số tài khoản ưu tiên từ status.info (investor_profile_management)
                account_number = partner_name
                try:
                    if partner:
                        status_info = request.env['status.info'].sudo().search([('partner_id', '=', partner.id)], limit=1)
                        if status_info and getattr(status_info, 'so_tk', False):
                            account_number = status_info.so_tk
                except Exception:
                    # Fallback giữ nguyên partner_name nếu không có module/record
                    account_number = partner_name

                # Get nationality from investor.profile if available
                nationality = ''
                try:
                    if partner:
                        investor_profile = request.env['investor.profile'].sudo().search([('partner_id', '=', partner.id)], limit=1)
                        if investor_profile and investor_profile.nationality:
                            nationality = investor_profile.nationality.name
                except Exception:
                    nationality = ''

                sheet.cell(row=i, column=1, value=i-1)  # STT
                sheet.cell(row=i, column=2, value=account_number)
                sheet.cell(row=i, column=3, value=user_name)
                sheet.cell(row=i, column=4, value='Trực tiếp')  # Default investor type
                sheet.cell(row=i, column=5, value=nationality)
                sheet.cell(row=i, column=6, value=record.units or 0)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to buffer
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Set response headers
            xlsx_headers = [
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="report_balance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'),
            ]
            
            return request.make_response(output.read(), headers=xlsx_headers)
            
        except Exception as e:
            print(f"Error in export_xlsx: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất XLSX</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    @http.route('/report-balance/export-csv', type='http', auth='user', website=True)
    def export_csv(self, **kw):
        """
        Route to export the report balance data as a CSV.
        """
        try:
            # Sử dụng portfolio.investment thay vì report.balance
            report_model = request.env['portfolio.investment'].sudo()
            
            # Get filters from URL parameters
            selected_product = kw.get('product', '')
            from_date = kw.get('from_date', '')
            to_date = kw.get('to_date', '')
            
            # Build domain based on filters
            domain = []
            if selected_product:
                domain.append(['fund_id.name', '=', selected_product])
            if from_date:
                domain.append(['create_date', '>=', from_date])
            if to_date:
                domain.append(['create_date', '<=', to_date])
            
            # Get all records for CSV export (no pagination)
            records = report_model.search(domain, order='id desc')
            
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Số tài khoản',
                'Nhà đầu tư',
                'Số điện thoại',
                'ĐKSH',
                'Email',
                'Loại NĐT',
                'Quỹ',
                'Chương trình',
                'Số CCQ'
            ])
            
            # Write data rows
            for record in records:
                # Get user/partner info
                user_name = record.user_id.name if record.user_id else ''
                partner = record.user_id.partner_id if record.user_id else False
                partner_name = partner.name if partner else ''

                # Fetch số tài khoản ưu tiên từ status.info (investor_profile_management)
                account_number = partner_name
                try:
                    if partner:
                        status_info = request.env['status.info'].sudo().search([('partner_id', '=', partner.id)], limit=1)
                        if status_info and getattr(status_info, 'so_tk', False):
                            account_number = status_info.so_tk
                except Exception:
                    # Fallback giữ nguyên partner_name nếu không có module/record
                    account_number = partner_name

                # Get fund info
                fund_name = record.fund_id.name if record.fund_id else ''
                fund_ticker = record.fund_id.ticker if record.fund_id else ''

                writer.writerow([
                    account_number,
                    user_name,
                    record.user_id.partner_id.phone if record.user_id and record.user_id.partner_id else '',
                    str(record.amount) if record.amount else '',
                    record.user_id.partner_id.email if record.user_id and record.user_id.partner_id else '',
                    'Trực tiếp',  # Default investor type
                    fund_name,
                    fund_ticker,
                    self._format_currency(record.units) if record.units else '0'
                ])
            
            csv_content = output.getvalue()
            output.close()
            
            # Set response headers
            csv_http_headers = [
                ('Content-Type', 'text/csv; charset=utf-8'),
                ('Content-Length', len(csv_content.encode('utf-8'))),
                ('Content-Disposition', f'attachment; filename="Bao_cao_so_du_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'),
            ]
            
            return request.make_response(csv_content, headers=csv_http_headers)
            
        except Exception as e:
            print(f"Error in export_csv: {str(e)}")
            import traceback
            traceback.print_exc()
            return request.make_response(
                f"<h1>Lỗi xuất CSV</h1><p>Chi tiết lỗi: {str(e)}</p>",
                headers=[('Content-Type', 'text/html')]
            )

    def _format_currency(self, amount):
        """Helper method to format currency"""
        if not amount:
            return '0'
        return f"{amount:,.0f}".replace(',', '.')

    def _format_date(self, date_value):
        """Helper method to format date"""
        if not date_value:
            return ''
        if isinstance(date_value, str):
            try:
                date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
            except:
                return date_value
        return date_value.strftime('%d/%m/%Y')

    def _get_investor_type_label(self, investor_type):
        """Helper method to get investor type label"""
        investor_types = {
            'truc_tiep': 'Trực tiếp',
            'ky_danh': 'Ký danh',
        }
        return investor_types.get(investor_type, investor_type)
